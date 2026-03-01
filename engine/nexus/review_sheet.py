"""Review Sheet — generates and imports Excel review sheets for human QA review.

Creates an openpyxl Excel workbook with formulas, dropdown validation, and
conditional formatting so that Q&A pairs can be reviewed, approved, or rejected
by a human before being imported into the Nexus cache.

Layout:
  Column A: Question
  Column B: Answer
  Column C: Consumer         (data validation dropdown)
  Column D: Priority         (1-5, conditional formatting: 5=green, 1=red)
  Column E: Category         (data validation dropdown)
  Column F: NLM_Rating       (ESSENTIAL/USEFUL/SKIP — pre-filled, editable dropdown)
  Column G: Include?         (formula: =IF(OR(F2="ESSENTIAL",F2="USEFUL"),"YES","REVIEW"))
  Column H: Duplicate?       (formula: =COUNTIF($A$1:A1,A2)>0)
  Column I: Notes            (free text for reviewer)

Row 1: frozen header, auto-filter on all columns.

Usage::

    from engine.nexus.review_sheet import get_review_sheet
    rs = get_review_sheet()

    # Generate from pipeline output
    path = rs.generate(approved_pairs, "data/qa_review_20260301.xlsx")

    # Import reviewed file back to Nexus
    from engine.nexus.client import get_nexus_client
    count = rs.import_reviewed(path, get_nexus_client())
"""
from __future__ import annotations

import logging
import threading
from pathlib import Path
from typing import Any, List, Optional

logger = logging.getLogger(__name__)

# ──── Schema Constants ────────────────────────────────────────────────────────

COLUMNS = [
    "Question",
    "Answer",
    "Consumer",
    "Priority",
    "Category",
    "NLM_Rating",
    "Include?",
    "Duplicate?",
    "Notes",
]

CONSUMERS = ["copilot", "agent", "governance", "developer", "news"]
CATEGORIES = [
    "architecture", "skills", "config", "testing",
    "nexus", "tools", "scenes", "general",
]
RATINGS = ["ESSENTIAL", "USEFUL", "SKIP"]

# Include formula (column G, row 2+)
# =IF(OR(F2="ESSENTIAL",F2="USEFUL"),"YES","REVIEW")
_INCLUDE_FORMULA = '=IF(OR(F{r}="ESSENTIAL",F{r}="USEFUL"),"YES","REVIEW")'

# Duplicate detection formula (column H, row 2+)
# =COUNTIF($A$1:A1,A2)>0
_DUPLICATE_FORMULA = "=COUNTIF($A$1:A{prev},A{r})>0"

# Column widths (characters)
_COL_WIDTHS = {
    "A": 60,   # Question
    "B": 80,   # Answer
    "C": 14,   # Consumer
    "D": 10,   # Priority
    "E": 16,   # Category
    "F": 14,   # NLM_Rating
    "G": 12,   # Include?
    "H": 12,   # Duplicate?
    "I": 30,   # Notes
}


# ──── Review Sheet ────────────────────────────────────────────────────────────

class ReviewSheet:
    """Generates and imports Excel review sheets for Nexus Q&A pairs.

    Uses openpyxl for Excel generation.  openpyxl is a lightweight pure-Python
    library — no Excel installation required.
    """

    def generate(
        self,
        pairs: List[Any],
        path: str,
    ) -> str:
        """Build and save an Excel review sheet.

        Args:
            pairs: List of CandidatePair (or dict with q, a, consumer,
                priority, category, rating fields).
            path: File path to save the xlsx file.

        Returns:
            The saved file path, or empty string on failure.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment,
                Font,
                PatternFill,
                Protection,
            )
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            logger.error(
                "openpyxl is required for review sheets: pip install openpyxl"
            )
            return ""

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "QA Review"

            # ── Header Row ──────────────────────────────────────────────────
            header_fill = PatternFill(
                start_color="2E4057", end_color="2E4057", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col_idx, col_name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Freeze header row, enable auto-filter
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:I{len(pairs) + 1}"

            # ── Column Widths ──────────────────────────────────────────────
            for col_letter, width in _COL_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width

            # Row height for header
            ws.row_dimensions[1].height = 20

            # ── Data Rows ──────────────────────────────────────────────────
            for row_idx, pair in enumerate(pairs, start=2):
                q, a, consumer, priority, category, rating = self._extract_fields(pair)

                ws.cell(row=row_idx, column=1, value=q)
                ws.cell(row=row_idx, column=2, value=a)
                ws.cell(row=row_idx, column=3, value=consumer)
                ws.cell(row=row_idx, column=4, value=priority)
                ws.cell(row=row_idx, column=5, value=category)
                ws.cell(row=row_idx, column=6, value=rating or "USEFUL")

                # Formula: Include?
                ws.cell(
                    row=row_idx,
                    column=7,
                    value=_INCLUDE_FORMULA.format(r=row_idx),
                )

                # Formula: Duplicate? (compare to all rows above)
                ws.cell(
                    row=row_idx,
                    column=8,
                    value=_DUPLICATE_FORMULA.format(r=row_idx, prev=row_idx - 1),
                )

                # Notes — blank for reviewer
                ws.cell(row=row_idx, column=9, value="")

                # Wrap text in Q and A columns
                ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
                ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)

            # ── Data Validation: Consumer dropdown ─────────────────────────
            if len(pairs) > 0:
                consumer_dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(CONSUMERS)}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                consumer_dv.sqref = f"C2:C{len(pairs) + 1}"
                ws.add_data_validation(consumer_dv)

            # ── Data Validation: Category dropdown ─────────────────────────
            if len(pairs) > 0:
                category_dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(CATEGORIES)}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                category_dv.sqref = f"E2:E{len(pairs) + 1}"
                ws.add_data_validation(category_dv)

            # ── Data Validation: Rating dropdown ───────────────────────────
            if len(pairs) > 0:
                rating_dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(RATINGS)}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                rating_dv.sqref = f"F2:F{len(pairs) + 1}"
                ws.add_data_validation(rating_dv)

            # ── Data Validation: Priority 1-5 ──────────────────────────────
            if len(pairs) > 0:
                priority_dv = DataValidation(
                    type="whole",
                    operator="between",
                    formula1="1",
                    formula2="5",
                    allow_blank=False,
                )
                priority_dv.sqref = f"D2:D{len(pairs) + 1}"
                ws.add_data_validation(priority_dv)

            # ── Conditional Formatting: Priority color scale ────────────────
            if len(pairs) > 0:
                priority_range = f"D2:D{len(pairs) + 1}"
                ws.conditional_formatting.add(
                    priority_range,
                    ColorScaleRule(
                        start_type="num", start_value=1, start_color="FF6B6B",  # red
                        mid_type="num", mid_value=3, mid_color="FFE66D",        # yellow
                        end_type="num", end_value=5, end_color="6BCB77",        # green
                    ),
                )

                # SKIP rows → grey background
                skip_fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                )
                ws.conditional_formatting.add(
                    f"A2:I{len(pairs) + 1}",
                    CellIsRule(
                        operator="equal",
                        formula=['"SKIP"'],
                        fill=skip_fill,
                    ),
                )

            # ── Save ────────────────────────────────────────────────────────
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(path)
            logger.info("Review sheet saved: %s (%d rows)", path, len(pairs))
            return path

        except Exception as exc:
            logger.error("Failed to generate review sheet: %s", exc, exc_info=True)
            return ""

    def import_reviewed(self, path: str, client: Any) -> int:
        """Import approved rows from a reviewed Excel file into Nexus.

        Reads all rows where Include? column == "YES" and stores them in the
        Nexus Q&A cache.

        Args:
            path: Path to the reviewed xlsx file.
            client: NexusClient instance.

        Returns:
            Number of pairs successfully stored in Nexus.
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl is required: pip install openpyxl")
            return 0

        if not Path(path).exists():
            logger.error("Review sheet not found: %s", path)
            return 0

        if not client or not client.is_available():
            logger.error("Nexus client unavailable")
            return 0

        stored = 0
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active

            # Find column indices from header row
            header = {
                cell.value: cell.column
                for cell in ws[1]
                if cell.value
            }

            col_q = header.get("Question", 1)
            col_a = header.get("Answer", 2)
            col_consumer = header.get("Consumer", 3)
            col_priority = header.get("Priority", 4)
            col_category = header.get("Category", 5)
            col_include = header.get("Include?", 7)

            for row in ws.iter_rows(min_row=2, values_only=True):
                include = str(row[col_include - 1] or "").strip().upper()
                if include != "YES":
                    continue

                q = str(row[col_q - 1] or "").strip()
                a = str(row[col_a - 1] or "").strip()
                if not q or not a:
                    continue

                consumer = str(row[col_consumer - 1] or "developer").lower().strip()
                try:
                    priority = int(row[col_priority - 1] or 3)
                except (ValueError, TypeError):
                    priority = 3
                category = str(row[col_category - 1] or "general").lower().strip()

                try:
                    client.add_qa(
                        question=q[:500],
                        answer=a[:2000],
                        category=category,
                        tags=[
                            "human-reviewed",
                            f"consumer-{consumer}",
                            f"priority-{priority}",
                        ],
                    )
                    stored += 1
                except Exception as exc:
                    logger.debug("Failed to store reviewed pair: %s", exc)

        except Exception as exc:
            logger.error("Failed to import reviewed sheet: %s", exc, exc_info=True)

        logger.info("Imported %d approved pairs from %s", stored, path)
        return stored

    # ── Internal ─────────────────────────────────────────────────────────────

    def _extract_fields(
        self, pair: Any
    ) -> tuple:
        """Extract (q, a, consumer, priority, category, rating) from a pair object or dict."""
        if isinstance(pair, dict):
            q = str(pair.get("q", pair.get("question", ""))).strip()
            a = str(pair.get("a", pair.get("answer", ""))).strip()
            consumer = str(pair.get("consumer", "developer")).lower()
            try:
                priority = int(pair.get("priority", 3))
            except (ValueError, TypeError):
                priority = 3
            category = str(pair.get("category", "general")).lower()
            rating = str(pair.get("rating", pair.get("nlm_rating", ""))).upper()
        else:
            # CandidatePair dataclass
            q = str(getattr(pair, "q", "")).strip()
            a = str(getattr(pair, "a", "")).strip()
            consumer = str(getattr(pair, "consumer", "developer")).lower()
            priority = int(getattr(pair, "priority", 3))
            category = str(getattr(pair, "category", "general")).lower()
            rating = str(getattr(pair, "rating", "")).upper()
        return q, a, consumer, max(1, min(5, priority)), category, rating


# ──── Singleton ───────────────────────────────────────────────────────────────

_sheet_instance: Optional[ReviewSheet] = None
_sheet_lock = threading.Lock()


def get_review_sheet() -> ReviewSheet:
    """Get the singleton ReviewSheet instance."""
    global _sheet_instance
    if _sheet_instance is None:
        with _sheet_lock:
            if _sheet_instance is None:
                _sheet_instance = ReviewSheet()
    return _sheet_instance
